"""
Eenvoudige tag-recommender op basis van een trainingsset in Excel.
Gebruikt een bag-of-words benadering met IDF-weging om per tag een score te berekenen.
"""
import logging
import math
import os
import re
from collections import Counter, defaultdict
from typing import Dict, List, Tuple

from openpyxl import load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.pipeline import make_pipeline

TOKEN_RE = re.compile(r"[A-Za-z0-9]+")


class TagRecommender:
    """Houdt een lichtgewicht vocabulaire per tag bij en kan suggesties genereren."""

    def __init__(self, training_path: str, allowed_tags: List[str] | None = None, additional_data_path: str | None = None):
        self.training_path = training_path
        self.additional_data_path = additional_data_path  # Bijv. werkbestand met al ingevulde tags
        self.allowed_tags = set(allowed_tags or [])
        self.tag_token_freq: defaultdict[str, Counter[str]] = defaultdict(Counter)
        self.token_doc_freq: Counter[str] = Counter()
        self.tag_totals: Counter[str] = Counter()
        self.total_docs = 0
        self.last_loaded_mtime: float | None = None
        self.last_additional_mtime: float | None = None
        self.model = None

    @staticmethod
    def _tokenize(text: str) -> List[str]:
        """Tokenizeer tekst en voeg samengestelde woorden toe."""
        # Basis tokenisatie
        basic_tokens = [match.group(0).lower() for match in TOKEN_RE.finditer(text or "")]
        
        # Voeg aanvullende tokens toe voor betere matching
        # Bijv. "jeugdlid" -> ook "jeugd"
        extra_tokens = []
        for token in basic_tokens:
            if "jeugd" in token:
                extra_tokens.append("jeugd")
            if "volwassenen" in token or "volwassen" in token:
                extra_tokens.append("volwassenen")
        
        return basic_tokens + extra_tokens

    def _reset(self) -> None:
        self.tag_token_freq.clear()
        self.token_doc_freq.clear()
        self.tag_totals.clear()
        self.total_docs = 0

    def _find_columns(self, header: List[str]) -> Tuple[int | None, List[int]]:
        """Zoek de kolommen voor tag en tekstvelden."""
        normalized = [str(col).strip().lower() for col in header]
        lookup = {name: idx for idx, name in enumerate(normalized)}

        tag_col = None
        for candidate in ("tag", "tags", "categorie", "category"):
            if candidate in lookup:
                tag_col = lookup[candidate]
                break

        text_cols = []
        for candidate in (
            "naam / omschrijving",
            "naam/omschrijving",
            "mededeling",
            "mededelingen",
            "omschrijving",
            "rekening",
            "tegenrekening",
            "mutatiesoort",
            "memo",
            "code",
            "description",
        ):
            if candidate in lookup:
                text_cols.append(lookup[candidate])

        if not text_cols:
            # Geen duidelijke tekstkolommen gevonden: gebruik alle kolommen behalve de tagkolom
            text_cols = [idx for idx in range(len(header)) if idx != tag_col]

        return tag_col, text_cols

    def _collect_dataset(self, path: str) -> List[tuple[str, str]]:
        """Lees een Excelbestand en verzamel (text, tag) voorbeelden."""
        samples: List[tuple[str, str]] = []
        wb = None
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            for sheet in wb.worksheets:
                first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
                if not first_row:
                    continue
                header = [str(val).strip() if val is not None else "" for val in first_row]
                normalized = [str(col).strip().lower() for col in header]
                lookup = {name: idx for idx, name in enumerate(normalized)}

                # Zoek tag kolom
                tag_col = None
                for candidate in ("tag", "tags", "categorie", "category"):
                    if candidate in lookup:
                        tag_col = lookup[candidate]
                        break
                if tag_col is None:
                    continue

                # Zoek tekstkolommen
                text_cols = []
                for candidate in (
                    "naam / omschrijving",
                    "naam/omschrijving",
                    "mededeling",
                    "mededelingen",
                    "omschrijving",
                    "rekening",
                    "tegenrekening",
                    "mutatiesoort",
                    "memo",
                    "code",
                    "description",
                ):
                    if candidate in lookup:
                        text_cols.append(lookup[candidate])
                if not text_cols:
                    text_cols = [idx for idx in range(len(header)) if idx != tag_col]

                # Zoek bedrag kolom (optioneel)
                amount_col = None
                for candidate in ("bedrag (eur)", "bedrag", "amount"):
                    if candidate in lookup:
                        amount_col = lookup[candidate]
                        break

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not row or len(row) <= tag_col:
                        continue
                    tag_val = str(row[tag_col] or "").strip()
                    if not tag_val:
                        continue
                    if self.allowed_tags and tag_val not in self.allowed_tags:
                        continue

                    parts: List[str] = []
                    for idx in text_cols:
                        if idx < len(row) and row[idx] not in (None, ""):
                            parts.append(str(row[idx]))

                    # Voeg bedrag als speciaal token toe
                    if amount_col is not None and amount_col < len(row) and row[amount_col] not in (None, ""):
                        try:
                            amount_val = float(str(row[amount_col]).replace(",", "."))
                            parts.append(f"AMT_{round(amount_val)}")
                        except (ValueError, TypeError):
                            pass

                    if not parts:
                        continue

                    combined = " ".join(parts)
                    samples.append((combined, tag_val))
        except Exception as exc:  # noqa: BLE001
            logging.error("Fout bij laden dataset uit %s: %s", path, exc)
        finally:
            if wb:
                wb.close()
        return samples

    def _process_heuristic_sample(self, text: str, tag: str) -> None:
        """Verwerk een sample voor heuristische benadering."""
        tokens = self._tokenize(text)
        for token in tokens:
            self.tag_token_freq[tag][token] += 1
            self.token_doc_freq[token] += 1
        self.tag_totals[tag] += 1
        self.total_docs += 1

    def load(self) -> bool:
        """Train het ML-model op trainingsdata + reeds getagde werkdata."""
        if not self.training_path or not os.path.exists(self.training_path):
            logging.warning("Trainingsbestand niet gevonden: %s", self.training_path)
            return False

        # Hertrain alleen als bronbestanden gewijzigd zijn
        mtimes = [os.path.getmtime(self.training_path)]
        if self.additional_data_path and os.path.exists(self.additional_data_path):
            mtimes.append(os.path.getmtime(self.additional_data_path))
        latest_mtime = max(mtimes)
        if self.last_loaded_mtime and latest_mtime <= self.last_loaded_mtime:
            return True

        self._reset()

        # Verzamel training samples
        samples = self._collect_dataset(self.training_path)
        if self.additional_data_path and os.path.exists(self.additional_data_path):
            samples += self._collect_dataset(self.additional_data_path)

        if not samples:
            logging.warning("Geen trainingsdata gevonden om model te trainen")
            return False

        texts, labels = zip(*samples)

        # Controleer aantal unieke klassen
        unique_classes = set(labels)
        if len(unique_classes) < 2:
            logging.warning(
                "Onvoldoende trainingsklassen (%d) voor ML model; gebruik heuristische benadering",
                len(unique_classes)
            )
            # Bouw heuristische tag-vocabulaire
            for text, label in samples:
                self._process_heuristic_sample(text, label)
            self.model = None  # Markeer dat heuristics gebruikt worden
            self.last_loaded_mtime = latest_mtime
            return True

        # ML pipeline: TF-IDF (1-2 grams) + Logistic Regression
        model = make_pipeline(
            TfidfVectorizer(ngram_range=(1, 2), min_df=1),
            LogisticRegression(max_iter=1000, n_jobs=1, multi_class="auto")
        )

        try:
            model.fit(texts, labels)
            self.model = model
            self.last_loaded_mtime = latest_mtime
            logging.info("ML model getraind met %d voorbeelden", len(samples))
            return True
        except ValueError as exc:
            logging.error("ML model training mislukt: %s; valt terug op heuristics", exc)
            # Fallback: bouw heuristische tag-vocabulaire
            for text, label in samples:
                self._process_heuristic_sample(text, label)
            self.model = None
            self.last_loaded_mtime = latest_mtime
            return True

    def recommend(self, transaction: Dict[str, str], top_k: int = 3) -> List[Dict[str, float | str]]:
        """Geef een lijst met tags en scores terug op basis van het ML-model of heuristics."""
        if not self.load():
            return []

        parts: List[str] = []
        for key in (
            "mededelingen",
            "omschrijving",
            "naam",
            "rekening",
            "tegenrekening",
            "mutatiesoort",
            "code",
            "memo",
        ):
            val = transaction.get(key)
            if val:
                parts.append(str(val))

        # Voeg bedrag toe als token
        bedrag_str = str(transaction.get("bedrag", "")).strip()
        try:
            bedrag = float(bedrag_str.replace(",", ".")) if bedrag_str else None
        except (ValueError, AttributeError):
            bedrag = None
        if bedrag is not None:
            parts.append(f"AMT_{round(bedrag)}")

        if not parts:
            return []

        text = " ".join(parts)

        # Probeer ML-model te gebruiken
        if hasattr(self, "model") and self.model is not None:
            try:
                proba = self.model.predict_proba([text])[0]
                classes = self.model.classes_
                paired = sorted(zip(classes, proba), key=lambda p: p[1], reverse=True)
                return [
                    {"tag": tag, "score": round(float(score), 4)}
                    for tag, score in paired[:top_k]
                ]
            except Exception as exc:  # noqa: BLE001
                logging.error("Fout bij ML aanbeveling: %s", exc)

        # Fallback: heuristische benadering
        tokens = self._tokenize(text)
        tag_scores: Dict[str, float] = {}

        for tag in self.tag_token_freq:
            score = 0.0
            for token in tokens:
                if token in self.tag_token_freq[tag]:
                    # TF-IDF-achtige scoring
                    tf = self.tag_token_freq[tag][token]
                    idf = math.log(self.total_docs / max(self.token_doc_freq[token], 1)) if self.total_docs > 0 else 0
                    score += tf * idf
            if score > 0:
                tag_scores[tag] = score

        # Filter op allowed_tags
        if self.allowed_tags:
            tag_scores = {tag: score for tag, score in tag_scores.items() if tag in self.allowed_tags}

        if not tag_scores:
            return []

        sorted_tags = sorted(tag_scores.items(), key=lambda p: p[1], reverse=True)
        return [
            {"tag": tag, "score": round(float(score), 4)}
            for tag, score in sorted_tags[:top_k]
        ]
